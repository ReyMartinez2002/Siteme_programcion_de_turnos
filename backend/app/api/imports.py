from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from sqlalchemy.orm import Session
from typing import Any
from app.database import get_db
from app.models.models import Rider, PanpayaStore, ExternalBrand
from openpyxl import load_workbook

router = APIRouter(prefix="/imports", tags=["imports"])


MAX_UPLOAD_SIZE = 5 * 1024 * 1024


def _normalize(value: Any) -> str:
    return str(value).strip().upper() if value is not None else ""


@router.post("/riders")
def import_riders(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        raise HTTPException(status_code=400, detail="Invalid file format")
    if file.size is not None and file.size > MAX_UPLOAD_SIZE:
        raise HTTPException(status_code=400, detail="File too large")
    workbook = load_workbook(file.file, data_only=True)
    sheet = workbook.active
    headers = [_normalize(cell.value) for cell in sheet[1]]
    required = ["NOMBRE", "TIPO"]
    for column in required:
        if column not in headers:
            raise HTTPException(status_code=400, detail=f"Missing column {column}")
    name_index = headers.index("NOMBRE")
    type_index = headers.index("TIPO")
    store_index = headers.index("SUCURSAL") if "SUCURSAL" in headers else None
    id_index = headers.index("CC") if "CC" in headers else None
    obs_index = headers.index("OBSERVACION") if "OBSERVACION" in headers else None
    created = 0
    updated = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        full_name = str(row[name_index]).strip() if row[name_index] else ""
        rider_type = str(row[type_index]).strip() if row[type_index] else ""
        store_code = (
            str(row[store_index]).strip()
            if store_index is not None and row[store_index]
            else ""
        )
        if not full_name or not rider_type:
            continue
        store = None
        if store_code:
            store = db.query(PanpayaStore).filter(PanpayaStore.code == store_code).first()
        existing = db.query(Rider).filter(Rider.full_name == full_name).first()
        payload = {
            "full_name": full_name,
            "rider_type": rider_type,
            "active": True,
            "store_id": store.id if store else None,
            "identification": str(row[id_index]).strip() if id_index is not None and row[id_index] else None,
            "observation": str(row[obs_index]).strip() if obs_index is not None and row[obs_index] else None,
        }
        if existing:
            for key, value in payload.items():
                setattr(existing, key, value)
            updated += 1
        else:
            db.add(Rider(**payload))
            created += 1
    db.commit()
    return {"created": created, "updated": updated}


@router.post("/stores")
def import_stores(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        raise HTTPException(status_code=400, detail="Invalid file format")
    if file.size is not None and file.size > MAX_UPLOAD_SIZE:
        raise HTTPException(status_code=400, detail="File too large")
    workbook = load_workbook(file.file, data_only=True)
    sheet = workbook.active
    headers = [_normalize(cell.value) for cell in sheet[1]]
    required = ["CODIGO", "NOMBRE"]
    for column in required:
        if column not in headers:
            raise HTTPException(status_code=400, detail=f"Missing column {column}")
    code_index = headers.index("CODIGO")
    name_index = headers.index("NOMBRE")
    zone_index = headers.index("ZONA") if "ZONA" in headers else None
    address_index = headers.index("DIRECCION") if "DIRECCION" in headers else None
    created = 0
    updated = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        code = str(row[code_index]).strip() if row[code_index] else ""
        name = str(row[name_index]).strip() if row[name_index] else ""
        if not code or not name:
            continue
        existing = db.query(PanpayaStore).filter(PanpayaStore.code == code).first()
        payload = {
            "code": code,
            "name": name,
            "zone": str(row[zone_index]).strip() if zone_index is not None and row[zone_index] else None,
            "address": str(row[address_index]).strip() if address_index is not None and row[address_index] else None,
        }
        if existing:
            for key, value in payload.items():
                setattr(existing, key, value)
            updated += 1
        else:
            db.add(PanpayaStore(**payload))
            created += 1
    db.commit()
    return {"created": created, "updated": updated}


@router.post("/brands")
def import_brands(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        raise HTTPException(status_code=400, detail="Invalid file format")
    if file.size is not None and file.size > MAX_UPLOAD_SIZE:
        raise HTTPException(status_code=400, detail="File too large")
    workbook = load_workbook(file.file, data_only=True)
    sheet = workbook.active
    headers = [_normalize(cell.value) for cell in sheet[1]]
    if "MARCA" not in headers:
        raise HTTPException(status_code=400, detail="Missing column MARCA")
    brand_index = headers.index("MARCA")
    created = 0
    updated = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name = str(row[brand_index]).strip() if row[brand_index] else ""
        if not name:
            continue
        existing = db.query(ExternalBrand).filter(ExternalBrand.name == name).first()
        if existing:
            updated += 1
            continue
        db.add(ExternalBrand(name=name))
        created += 1
    db.commit()
    return {"created": created, "updated": updated}
